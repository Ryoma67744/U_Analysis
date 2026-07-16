"""annotation_inspect.py - 登録済みデータの「化合物名（注釈）の有無」を、生データ本体を開かずに要約する。

生データ（.parquet/.txt）は >1GB で手元で開けないため、変換時に生成される小さな副産物だけを読む:
  - TIMS/SCiLS: サイドカー `<BASE>_feature_annotations.parquet`（1 feature 1 行・数 KB）。
    peak-list を検出したときのみ生成されるため「存在＝化合物名あり」。無ければ本体 parquet の
    フッタ schema メタ `b"peak_list"` の有無で判定（フッタのみ数 KB 読取）。
  - DESI: サイドカー無し。化合物名は正規化 `.txt` の化合物名行（named 形式は 3 行目＝index2）、
    または named 形式 csv/xlsx の 1 行目ヘッダ（`x,y,<化合物名>_...`）から取得。

いずれも生データ本体は読まない（KB 級のサイドカー/フッタ/先頭数行のみ）。Dash 非依存の純ロジック。
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Optional

import pandas as pd

from app.utils.deg_utils import is_meaningful_annotation

logger = logging.getLogger(__name__)

_SIDE_GLOB = "*_feature_annotations.parquet"
_NUMERIC_RE = re.compile(r"^-?\d+\.?\d*$")
_EXAMPLE_COLS = ("mz", "display_name", "compound", "adduct", "formula")


# --------------------------------------------------------------------------
# 判定ユーティリティ
# --------------------------------------------------------------------------
def _is_real_compound(c) -> bool:
    """空・数値のみ・'No DB hit' を除いた「本物の化合物名」か。"""
    if not isinstance(c, str):
        return False
    s = c.strip()
    if not s or s.lower() == "no db hit":
        return False
    return is_meaningful_annotation(s, "")


def _candidate_dirs(sub: Optional[dict]) -> list:
    """sub から探索対象フォルダ（data_folder / last_result_dir / output_dir）を集める。"""
    dirs = []
    seen = set()
    for key in ("data_folder", "last_result_dir", "output_dir"):
        v = (sub or {}).get(key)
        if v and str(v) not in seen:
            seen.add(str(v))
            dirs.append(Path(v))
    return dirs


def _is_desi(sub: Optional[dict]) -> bool:
    return "DESI" in str((sub or {}).get("ms_instrument", "") or "").upper()


# --------------------------------------------------------------------------
# TIMS/SCiLS: サイドカー・フッタメタ
# --------------------------------------------------------------------------
def find_annotation_sidecar(dirs) -> Optional[Path]:
    """dirs（と各直下サブフォルダ）から `*_feature_annotations.parquet` を探す。

    サイドカーは本体 parquet と同じフォルダ（organize モードでは `<BASE>_Transform/`）に置かれ、
    解析時に output_dir/last_result_dir へコピーされる。親フォルダは走査しない（別データセットの
    サイドカーを誤検出しないため）。
    """
    seen = set()
    for d in dirs:
        try:
            if d is None or str(d) in seen or not d.is_dir():
                continue
        except OSError:
            continue
        seen.add(str(d))
        hits = sorted(d.glob(_SIDE_GLOB))
        if hits:
            return hits[0]
        sub_hits = sorted(d.glob("*/" + _SIDE_GLOB))
        if sub_hits:
            return sub_hits[0]
    return None


_PEAKLIST_SCAN_CAP = 50


def _main_parquet_peaklist(dirs) -> Optional[str]:
    """本体 parquet のフッタ schema メタに `b"peak_list"` があればその値（ファイル名）を返す。

    フッタ（数 KB）のみ読む。サイドカーが欠けているが本体には peak-list メタがある場合の保険。
    結果フォルダに parquet が大量にある/ディスクが遅いと待ちが伸びるため、走査は
    `_PEAKLIST_SCAN_CAP` 件で打ち切る（初ヒットで即 return する挙動は維持）。
    """
    try:
        import pyarrow.parquet as pq
    except Exception:
        return None
    scanned = 0
    for d in dirs:
        try:
            if d is None or not Path(d).is_dir():
                continue
        except OSError:
            continue
        parquets = sorted(set(Path(d).glob("*.parquet")) | set(Path(d).glob("*/*.parquet")))
        for pth in parquets:
            if pth.name.endswith("_feature_annotations.parquet"):
                continue
            if scanned >= _PEAKLIST_SCAN_CAP:
                logger.info("peak_list メタ走査を %d 件で打ち切り", _PEAKLIST_SCAN_CAP)
                return None
            scanned += 1
            try:
                md = pq.read_schema(str(pth)).metadata or {}
            except Exception:
                continue
            if b"peak_list" in md:
                try:
                    return md[b"peak_list"].decode("utf-8")
                except Exception:
                    return pth.name
    return None


# --------------------------------------------------------------------------
# DESI: .txt 化合物名行 / named ヘッダ
# --------------------------------------------------------------------------
def _desi_tokens_from_txt(pth: Path) -> Optional[list]:
    """正規化 .txt の先頭 5 行から化合物名トークンを取り出す（無ければ None）。

    従来形式は m/z のみ（行4=数値）。named 形式は行3(index2)=化合物名・行4=空。
    非数値トークンが主体の行を化合物名行とみなす。
    """
    try:
        with open(pth, "r", encoding="utf-8", errors="replace") as fh:
            lines = [fh.readline() for _ in range(5)]
    except Exception:
        return None
    for idx in (2, 3):  # 行3・行4 を確認
        if idx >= len(lines):
            continue
        toks = [p.strip() for p in lines[idx].split("\t") if p.strip()]
        if not toks:
            continue
        nonnum = [t for t in toks if not _NUMERIC_RE.match(t)]
        if nonnum and len(nonnum) >= max(1, len(toks) // 2):
            return nonnum
    return None


def _xlsx_header_fast(pth: Path) -> list:
    """xlsx の先頭シート 1 行目だけを openpyxl の read_only で読む（全体パースを避ける）。

    `pd.read_excel(nrows=0)` は openpyxl でブック全体を読み込むため大きい xlsx で非常に遅い。
    ヘッダ行の判定に必要なのは 1 行目だけなので、遅延読取でその 1 行のみ取得する。
    空セルは "" にして pandas 版（`Unnamed: N`）と同様に下流でスキップされるようにする。
    """
    import openpyxl

    wb = openpyxl.load_workbook(pth, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    finally:
        wb.close()
    return ["" if c is None else c for c in first]


def _desi_tokens_from_named_header(pth: Path) -> Optional[list]:
    """named 形式 csv/xlsx の 1 行目（`x,y,<化合物名>_...`）から化合物名を取り出す（無ければ None）。"""
    try:
        if pth.suffix.lower() == ".csv":
            header = list(pd.read_csv(pth, nrows=0).columns)
        else:
            header = _xlsx_header_fast(pth)
    except Exception:
        return None
    if len(header) < 3:
        return None
    if str(header[0]).strip().lower() != "x" or str(header[1]).strip().lower() != "y":
        return None
    toks = []
    for h in header[2:]:
        s = str(h).strip()
        if not s or s.lower().startswith("unnamed"):
            continue
        name = s.split("_", 1)[0]
        if name and not _NUMERIC_RE.match(name):
            toks.append(name)
    return toks or None


def _desi_compound_tokens(dirs):
    """DESI の .txt / named csv・xlsx から化合物名トークンを取り出す。

    Returns: (tokens:list[str], source_file:str) または (None, None)。
    """
    for d in dirs:
        try:
            if d is None or not Path(d).is_dir():
                continue
        except OSError:
            continue
        txts = sorted(set(Path(d).glob("*.txt")) | set(Path(d).glob("*/*.txt")))
        for pth in txts:
            toks = _desi_tokens_from_txt(pth)
            if toks:
                return toks, str(pth)
        for ext in ("*.csv", "*.xlsx"):
            files = sorted(set(Path(d).glob(ext)) | set(Path(d).glob("*/" + ext)))
            for pth in files:
                toks = _desi_tokens_from_named_header(pth)
                if toks:
                    return toks, str(pth)
    return None, None


# --------------------------------------------------------------------------
# 公開 API
# --------------------------------------------------------------------------
def has_compound_names(sub: Optional[dict]) -> bool:
    """バッジ用の安価チェック（parquet 本体は読まない）。"""
    dirs = _candidate_dirs(sub)
    if not dirs:
        return False
    if _is_desi(sub):
        toks, _ = _desi_compound_tokens(dirs)
        return bool(toks)
    # 既定は TIMS/SCiLS: サイドカー存在が定義的シグナル（本体は開かない）
    if find_annotation_sidecar(dirs) is not None:
        return True
    # instrument 未指定なら DESI ヘッダも一応拾う
    if not (sub or {}).get("ms_instrument"):
        toks, _ = _desi_compound_tokens(dirs)
        return bool(toks)
    return False


def _empty_result() -> dict:
    return {
        "status": "none",          # "annotated" | "none" | "unknown"
        "n_annotated": 0,
        "n_total": 0,
        "coverage_pct": None,
        "examples": [],            # [{mz, display_name, compound, adduct, formula}]
        "source_file": None,
        "source_kind": None,       # "sidecar" | "parquet_meta" | "desi_header"
        "note": "",
    }


def _summarize_sidecar(df: "pd.DataFrame", sidecar: Path, max_examples: int,
                       result: dict) -> dict:
    result["source_kind"] = "sidecar"
    result["source_file"] = str(sidecar)
    total = int(len(df))
    result["n_total"] = total
    if total == 0 or "compound" not in df.columns:
        result["status"] = "none"
        result["coverage_pct"] = 0.0 if total else None
        result["note"] = "サイドカーはありますが化合物名の行がありません。"
        return result
    mask = df["compound"].apply(_is_real_compound)
    n_ann = int(mask.sum())
    result["n_annotated"] = n_ann
    result["coverage_pct"] = round(100.0 * n_ann / total, 1) if total else None
    result["status"] = "annotated" if n_ann > 0 else "none"
    cols = [c for c in _EXAMPLE_COLS if c in df.columns]
    if n_ann:
        result["examples"] = df.loc[mask, cols].head(max_examples).to_dict("records")
    else:
        result["note"] = "化合物名は 0 件でした（No DB hit / m/z のみ）。"
    return result


def _inspect_desi(dirs, max_examples: int, result: dict) -> dict:
    toks, src = _desi_compound_tokens(dirs)
    result["source_kind"] = "desi_header"
    if not toks:
        result["status"] = "none"
        result["note"] = "DESI データに化合物名ヘッダは見つかりませんでした（m/z のみ）。"
        return result
    result["source_file"] = src
    result["n_total"] = len(toks)
    result["n_annotated"] = len(toks)
    result["coverage_pct"] = 100.0
    result["status"] = "annotated"
    result["examples"] = [{"compound": t} for t in toks[:max_examples]]
    return result


# --------------------------------------------------------------------------
# キャッシュ（同一状態の再オープンを即時化）
# --------------------------------------------------------------------------
_INSPECT_CACHE: dict = {}
_INSPECT_CACHE_MAX = 256


def _cache_signature(sub: Optional[dict], max_examples: int) -> str:
    """候補フォルダのパス＋mtime を元にした軽い署名。フォルダ内のファイル増減
    （サイドカー生成/削除・変換のやり直し等）で mtime が変わりキャッシュが無効化される。
    """
    parts = [
        str((sub or {}).get("id", "")),
        str((sub or {}).get("ms_instrument", "")),
        str(max_examples),
    ]
    for d in _candidate_dirs(sub):
        try:
            parts.append(f"{d}:{d.stat().st_mtime_ns}")
        except OSError:
            parts.append(f"{d}:NA")
    return "|".join(parts)


def inspect_annotations(sub: Optional[dict], max_examples: int = 200) -> dict:
    """モーダル用の詳細サマリを返す（生データ本体は読まない）。署名ベースでメモ化する。"""
    sig = _cache_signature(sub, max_examples)
    cached = _INSPECT_CACHE.get(sig)
    if cached is not None:
        return dict(cached)  # 呼び出し側の変更から守るため浅いコピーを返す
    result = _inspect_annotations_uncached(sub, max_examples)
    if len(_INSPECT_CACHE) >= _INSPECT_CACHE_MAX:
        _INSPECT_CACHE.clear()  # 上限超過時は単純に一掃（LRU は不要）
    _INSPECT_CACHE[sig] = dict(result)
    return result


def _inspect_annotations_uncached(sub: Optional[dict], max_examples: int = 200) -> dict:
    result = _empty_result()
    dirs = _candidate_dirs(sub)
    if not dirs:
        result["status"] = "unknown"
        result["note"] = "データフォルダが未設定です。"
        return result

    if _is_desi(sub):
        return _inspect_desi(dirs, max_examples, result)

    # TIMS/SCiLS: サイドカー優先
    sidecar = find_annotation_sidecar(dirs)
    if sidecar is not None:
        try:
            df = pd.read_parquet(sidecar)
        except Exception as e:
            result["status"] = "unknown"
            result["source_file"] = str(sidecar)
            result["note"] = f"サイドカーの読取に失敗しました: {e}"
            return result
        return _summarize_sidecar(df, sidecar, max_examples, result)

    # フッタ b"peak_list" フォールバック（サイドカー欠如時）
    pl = _main_parquet_peaklist(dirs)
    if pl:
        result["status"] = "annotated"
        result["source_kind"] = "parquet_meta"
        result["source_file"] = pl
        result["note"] = ("本体 parquet に peak-list メタ（化合物名あり）を検出しました。"
                          "サイドカーが無いため件数一覧は表示できません。")
        return result

    # instrument 未指定なら DESI も試す
    if not (sub or {}).get("ms_instrument"):
        desi = _inspect_desi(dirs, max_examples, _empty_result())
        if desi["status"] == "annotated":
            return desi

    result["note"] = "化合物名（注釈）は含まれていません（m/z のみ）。"
    return result
