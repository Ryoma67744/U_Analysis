# =============================================================================
# MSI Analysis Application - DESI Excel/CSV → 正規 .txt 変換器
# =============================================================================
#
# DESI の登録データは従来タブ区切りの `.txt`（先頭4行ヘッダ + ピクセルデータ）
# のみ対応していた。本モジュールは「中身のレイアウトは従来と完全に同一のまま、
# 容器（ファイル形式）だけ Excel(.xlsx/.xls) / CSV(.csv) で登録したい」という
# 要件に対し、入力を内部で正規 `.txt`（タブ区切り・同一レイアウト）へ変換する。
#
# これにより R 本解析 (read_desi_data) も Python の各リーダー
# (list_msi_files / read_desi_roi_list / validate_msi_file) も変更不要で、
# 変換後の `.txt` を透過的に処理できる。
#
# 設計方針:
#   - 行・セル構造は一切正規化せず「そのまま」タブ区切りで再出力する
#     (ragged 行・空1行目を保持。R fread(fill=TRUE) / ROI 検出が前提とするため)。
#   - CSV はクオート対応パース → m/z 小数は文字列のまま保持（桁落ちゼロ）。
#   - Excel はセルを原表現で文字列化（指数表記・桁落ちを抑止）。
#   - 冪等: 既存 `.txt` が新しければ再変換しない (mtime 比較 + 変換マーカー)。
#   - 循環 import 回避のため data_manager は import しない。
# =============================================================================

import csv
import logging
import os
from pathlib import Path
from typing import Optional

logger = logging.getLogger("msi.desi_converter")

# DESI 生データとして受け付ける「変換元」拡張子（.txt は変換対象外）
DESI_SRC_EXTS = (".csv", ".xlsx", ".xls")

# Excel 数値セルの文字列化フォーマット（指数表記を抑止しつつ有効桁を確保）
_FLOAT_FMT = ".10g"


# ---------------------------------------------------------------------------
# 変換元ファイルの探索
# ---------------------------------------------------------------------------

def find_desi_source(data_folder, stem: str) -> Optional[Path]:
    """data_folder 内の `<stem>` に対応する変換元 (.csv/.xlsx/.xls) を返す。

    複数形式が同居する場合は DESI_SRC_EXTS の順 (.csv → .xlsx → .xls) で優先。
    見つからなければ None。
    """
    folder = Path(data_folder)
    if not folder.is_dir():
        return None
    for ext in DESI_SRC_EXTS:
        p = folder / f"{stem}{ext}"
        if p.is_file():
            return p
    return None


def list_desi_convertibles(data_folder) -> dict:
    """対応する `.txt` が存在しない `<stem>` の変換元を {stem: src_path} で返す。"""
    folder = Path(data_folder)
    if not folder.is_dir():
        return {}
    txt_stems = {f.stem for f in folder.glob("*.txt")}
    candidate_stems = {
        f.stem for f in folder.iterdir()
        if f.is_file() and f.suffix.lower() in DESI_SRC_EXTS
    } - txt_stems
    out = {}
    for stem in sorted(candidate_stems):
        src = find_desi_source(folder, stem)
        if src is not None:
            out[stem] = src
    return out


# ---------------------------------------------------------------------------
# 純粋変換: 1 ファイル (.csv/.xlsx/.xls) → 正規 .txt
# ---------------------------------------------------------------------------

def _sniff_delimiter(sample_text: str) -> str:
    """CSV の区切り文字を冒頭テキストの頻度から推定する。

    DESI データ行は列数が多いため、真の区切り文字が支配的になる。
    小数点はピリオド前提（従来 .txt と同一）なので、カンマ=区切りが既定。
    """
    comma = sample_text.count(",")
    semi = sample_text.count(";")
    tab = sample_text.count("\t")
    if tab > 0 and tab >= comma and tab >= semi:
        return "\t"
    if semi > comma:
        return ";"
    return ","


def _read_csv_rows(path: Path) -> list:
    """CSV を行ごとの list[str] として読む（クオート対応・空行は []）。"""
    with open(path, "r", encoding="utf-8", errors="replace", newline="") as f:
        head = []
        for _ in range(10):
            line = f.readline()
            if not line:
                break
            head.append(line)
    delim = _sniff_delimiter("".join(head))

    rows = []
    with open(path, "r", encoding="utf-8", errors="replace", newline="") as f:
        for row in csv.reader(f, delimiter=delim):
            rows.append([("" if c is None else str(c)) for c in row])
    return rows


def _cell_to_str(v) -> str:
    """Excel セル値を、桁落ち・指数表記を避けつつ文字列化する。"""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        if v != v:  # NaN（pandas 経由の空セル）
            return ""
        if v.is_integer():
            return str(int(v))
        return format(v, _FLOAT_FMT)
    return str(v)


def _read_xlsx_rows(path: Path) -> list:
    """.xlsx/.xlsm を openpyxl で直読し、行ごとの list[str] を返す。"""
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover - 環境依存
        raise RuntimeError(
            "Excel(.xlsx) の読み込みに openpyxl が必要です。"
        ) from e

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        if len(wb.worksheets) > 1:
            logger.warning(
                "Excel に複数シートがあります。先頭シート '%s' のみ使用: %s",
                ws.title, path.name,
            )
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [_cell_to_str(v) for v in row]
            # 末尾の空セルをトリム（元 .txt の「末尾0省略」レイアウトに合わせる）
            while cells and cells[-1] == "":
                cells.pop()
            rows.append(cells)
        return rows
    finally:
        wb.close()


def _read_xls_rows(path: Path) -> list:
    """.xls (旧形式) を pandas 経由で読む（xlrd 等が必要な場合あり）。"""
    try:
        import pandas as pd
    except ImportError as e:  # pragma: no cover - 環境依存
        raise RuntimeError(".xls の読み込みに pandas が必要です。") from e
    try:
        df = pd.read_excel(str(path), sheet_name=0, header=None, dtype=object)
    except Exception as e:
        raise RuntimeError(
            f".xls の読み込みに失敗しました（{e}）。.xlsx 形式での保存を推奨します。"
        ) from e
    rows = []
    for _, r in df.iterrows():
        cells = [_cell_to_str(v) for v in r.tolist()]
        while cells and cells[-1] == "":
            cells.pop()
        rows.append(cells)
    return rows


# ---------------------------------------------------------------------------
# 新形式 (1行ヘッダ・列名=化合物名) の検出と「正規4行ヘッダ」への組み替え
# ---------------------------------------------------------------------------

def _cell_is_numeric(s) -> bool:
    """セル文字列が数値か（空は数値扱い=欠損として ROI 判定に含めない）。"""
    s = (s or "").strip()
    if s == "":
        return True
    try:
        float(s)
        return True
    except ValueError:
        return False


def _is_named_format(rows) -> bool:
    """新形式か判定。先頭行の1・2セル目が x, y (大小無視) なら新形式。

    従来形式は先頭行が空 (1セル目が x でない) のため、先頭行だけで排他的に分岐できる。
    """
    if not rows or not rows[0]:
        return False
    header = rows[0]
    if len(header) < 3:
        return False
    c0 = (header[0] or "").strip().lower()
    c1 = (header[1] or "").strip().lower()
    return c0 == "x" and c1 == "y"


def _reshape_named_format(rows) -> list:
    """新形式 [x, y, 化合物名_情報.., (末尾ROIラベル)] を従来の正規4行ヘッダに組み替える。

    出力行:
      行1: 空
      行2: '','','' + '1'..'N'    (代謝物番号; R では未使用だが整合のため)
      行3: '','','' + 化合物名1..N (= 特徴量名。各列名の最初の '_' より前)
      行4: 空                       (Q3 が無い → R 側で pre(化合物名) をそのまま名前に採用)
      行5+: 連番ID, x, y, 強度1..N, [ROIラベル]
    """
    header = list(rows[0])
    while header and (header[-1] or "").strip() == "":
        header.pop()  # 末尾の空ヘッダ列を除去
    n_cols = len(header)
    data_rows = rows[1:]

    # 末尾列が「非数値主体」なら ROI ラベル列とみなす（領域別解析用・任意）
    roi_ci = None
    last_ci = n_cols - 1
    if last_ci >= 2:
        vals = [r[last_ci] for r in data_rows
                if last_ci < len(r) and (r[last_ci] or "").strip() != ""]
        if vals and sum(1 for v in vals if not _cell_is_numeric(v)) > len(vals) * 0.5:
            roi_ci = last_ci

    feature_cols = list(range(2, last_ci if roi_ci is not None else n_cols))

    # 化合物名 = 列名の最初の '_' より前 (例 Acetylcholine_15_10 -> Acetylcholine)
    feature_names = []
    for ci in feature_cols:
        h = (header[ci] or "").strip()
        feature_names.append(h.split("_", 1)[0] if h else h)

    n = len(feature_names)
    out = [
        [],
        ["", "", ""] + [str(i + 1) for i in range(n)],
        ["", "", ""] + feature_names,
        [],
    ]
    spot_id = 0
    for r in data_rows:
        if all((c or "").strip() == "" for c in r):
            continue  # 完全な空行はスキップ
        spot_id += 1
        x = r[0] if len(r) > 0 else ""
        y = r[1] if len(r) > 1 else ""
        intensities = [r[ci] if ci < len(r) else "" for ci in feature_cols]
        row_out = [str(spot_id), x, y] + intensities
        if roi_ci is not None:
            row_out.append(r[roi_ci] if roi_ci < len(r) else "")
        out.append(row_out)
    return out


def _maybe_reshape_named_format(rows) -> list:
    """新形式なら組み替え、そうでなければ（従来形式）そのまま返す。"""
    if _is_named_format(rows):
        logger.info(
            "DESI 新形式(1行ヘッダ・列名=化合物名)を検出 → 正規レイアウトに組み替え"
        )
        return _reshape_named_format(rows)
    return rows


def convert_desi_to_txt(src_path, dst_txt) -> Path:
    """単一の .csv/.xlsx/.xls を正規 .txt（タブ区切り・同一レイアウト）に変換する。

    行・セル構造はそのまま保持し、区切りのみタブに統一して書き出す。
    一時ファイル経由 + os.replace でアトミックに置換する。
    """
    src_path = Path(src_path)
    dst_txt = Path(dst_txt)
    ext = src_path.suffix.lower()

    if ext in (".xlsx", ".xlsm"):
        rows = _read_xlsx_rows(src_path)
    elif ext == ".xls":
        rows = _read_xls_rows(src_path)
    else:  # .csv / .tsv / その他テキスト
        rows = _read_csv_rows(src_path)

    # 新形式 (1行ヘッダ・列名=化合物名) なら従来の正規レイアウトへ組み替える。
    # 従来形式はそのまま (レイアウト保持の passthrough)。
    rows = _maybe_reshape_named_format(rows)

    dst_txt.parent.mkdir(parents=True, exist_ok=True)
    tmp = dst_txt.parent / f"{dst_txt.name}.{os.getpid()}.tmp"
    with open(tmp, "w", encoding="utf-8", newline="") as f:
        for row in rows:
            f.write("\t".join(row))
            f.write("\n")
    os.replace(tmp, dst_txt)
    logger.info("DESI 入力を正規化: %s → %s", src_path.name, dst_txt.name)
    return dst_txt


def normalize_desi_txt(txt_path) -> bool:
    """既存の `.txt` が新形式(1行ヘッダ・列名=化合物名)なら正規レイアウトに組み替えて
    その場で上書きする。従来形式(先頭空行)やファイル無しは何もしない。

    登録/アップロードの過程で `.csv` が `.txt` 化されてフォルダに入った場合など、
    変換器(.csv/.xlsx)経路を通らずに新形式 .txt が置かれるケースを救済する。
    冪等: 組み替え後は先頭行が空になり、再読込で `_is_named_format=False` となるため
    二重組み替えされない。

    Returns: 組み替えを行ったら True。
    """
    txt_path = Path(txt_path)
    if not txt_path.is_file():
        return False
    try:
        rows = _read_csv_rows(txt_path)  # 区切り自動判定 (tab/カンマ両対応)
    except Exception as e:
        logger.warning("DESI .txt 読み込みに失敗 (%s): %s", txt_path.name, e)
        return False
    if not _is_named_format(rows):
        return False  # 従来形式 → 触らない

    out = _reshape_named_format(rows)
    tmp = txt_path.parent / f"{txt_path.name}.{os.getpid()}.tmp"
    with open(tmp, "w", encoding="utf-8", newline="") as f:
        for row in out:
            f.write("\t".join(row))
            f.write("\n")
    os.replace(tmp, txt_path)
    logger.info("DESI 新形式の .txt を正規レイアウトに組み替え: %s", txt_path.name)
    return True


# ---------------------------------------------------------------------------
# 冪等変換 / 書込先判定（read-only マウント対策）
# ---------------------------------------------------------------------------

def _conv_marker(folder, stem: str) -> Path:
    """自動変換で生成した .txt であることを示す隠しマーカーのパス。

    ユーザーが手で置いた本物の .txt（マーカー無し）を誤って上書きしないために使う。
    """
    return Path(folder) / f".{stem}.desi_converted"


def _is_writable(folder: Path) -> bool:
    """folder に書き込み可能かを実テストで判定する（read-only マウント検知）。"""
    try:
        t = folder / f".write_test_{os.getpid()}.tmp"
        t.touch()
        t.unlink()
        return True
    except OSError:
        return False


def _staging_dir(folder: Path) -> Path:
    """read-only な data_folder 用に、書込可能な staging ディレクトリを返す。"""
    import hashlib

    from app.config import OUTPUT_DATA_DIR

    h = hashlib.sha1(str(Path(folder).resolve()).encode("utf-8")).hexdigest()[:12]
    d = Path(OUTPUT_DATA_DIR) / "_desi_staging" / h
    d.mkdir(parents=True, exist_ok=True)
    return d


def ensure_desi_txt(stem: str, data_folder, *, dest_dir=None,
                    force: bool = False) -> Optional[Path]:
    """`<stem>.txt` を保証して返す（必要なら .csv/.xlsx から変換）。

    - 変換元が無ければ、既存の `<stem>.txt` を返す（無ければ None）。
    - ユーザーが手で置いた `.txt`（変換マーカー無し）は上書きせず優先する。
    - dest_dir=None: data_folder が書込可能ならそこへ、読取専用なら staging へ出力。
    - dest_dir 指定: そのディレクトリへ出力（呼び出し側で staging を管理する場合）。
    - 冪等: 変換済み `.txt` が変換元より新しければ再変換しない。
    """
    folder = Path(data_folder)
    src = find_desi_source(folder, stem)
    real_txt = folder / f"{stem}.txt"

    # 変換元が無い → 既存 .txt をそのまま返す
    if src is None:
        return real_txt if real_txt.is_file() else None

    # ユーザー手置きの本物 .txt（マーカー無し）は最優先・上書きしない
    if (real_txt.is_file() and not _conv_marker(folder, stem).exists()
            and not force):
        return real_txt

    # 出力先の決定
    if dest_dir is not None:
        dest = Path(dest_dir)
    else:
        dest = folder if _is_writable(folder) else _staging_dir(folder)
    dest.mkdir(parents=True, exist_ok=True)

    target_txt = dest / f"{stem}.txt"
    marker = _conv_marker(dest, stem)

    # 既存の変換結果が新しければ再変換しない
    if (not force and target_txt.is_file() and marker.exists()
            and target_txt.stat().st_mtime >= src.stat().st_mtime):
        return target_txt

    convert_desi_to_txt(src, target_txt)
    try:
        marker.touch()
    except OSError:
        pass
    return target_txt


def prepare_desi_data_folder(data_folder, sample_names) -> str:
    """選択サンプル全件の `<stem>.txt` を 1 フォルダに揃え、R に渡すパスを返す。

    - data_folder が書込可能: その場で変換し、data_folder をそのまま返す。
    - 読取専用: staging に「ユーザー .txt のコピー」＋「変換 .txt」を集約し、
      staging パスを返す（R は 1 フォルダのみ参照するため）。
    解析起動前 (generate_v8_config) からの単一フックとして使う。
    """
    folder = Path(data_folder)
    if not folder.is_dir():
        return str(folder)

    if _is_writable(folder):
        for stem in (sample_names or []):
            try:
                ensure_desi_txt(stem, folder, dest_dir=folder)
                # .csv/.xlsx 経由でなく、新形式 .txt が直接置かれた場合も組み替える
                normalize_desi_txt(folder / f"{stem}.txt")
            except Exception as e:
                logger.warning("DESI 変換に失敗 (sample=%s): %s", stem, e)
        return str(folder)

    # --- read-only: staging に集約 ---
    import shutil

    staging = _staging_dir(folder)
    for stem in (sample_names or []):
        real_txt = folder / f"{stem}.txt"
        dst_txt = staging / f"{stem}.txt"
        try:
            if real_txt.is_file() and not _conv_marker(folder, stem).exists():
                # ユーザー手置きの .txt → staging にコピー
                if (not dst_txt.is_file()
                        or dst_txt.stat().st_mtime < real_txt.stat().st_mtime):
                    shutil.copy2(real_txt, dst_txt)
            elif find_desi_source(folder, stem) is not None:
                ensure_desi_txt(stem, folder, dest_dir=staging)
            elif real_txt.is_file():
                if not dst_txt.is_file():
                    shutil.copy2(real_txt, dst_txt)
            # staging 上の .txt が新形式ならここで組み替える
            normalize_desi_txt(dst_txt)
        except Exception as e:
            logger.warning("DESI staging 準備に失敗 (sample=%s): %s", stem, e)
    logger.info("DESI 入力を staging に集約 (read-only 元フォルダ): %s", staging)
    return str(staging)
